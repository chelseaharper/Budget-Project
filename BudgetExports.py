from BudgetCalculations import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def create_cash_flow(start, end, account, initial=0):
    wb = Workbook()
    dest_filename = "Test_File.xlsx"
    ws = wb.active
    ws.title = "Cash Flow"
    
    ws['A1'] = "Date"
    ws['B1'] = "Description"
    ws['C1'] = "Amount"
    ws['D1'] = "Account Balance"
    
    ws['A2'] = start
    ws['B2'] = "Initial Balance"
    ws['C2'] = initial
    ws['D2'] = initial

    data = account.ledger.coalate_transactions(start, end, account)
    for row in range(1, len(data)):
        for col in range(0, 3):
            _ = ws.cell(column = col, row = row, value = "")
    for i in data:
        pass
    wb.save(filename = dest_filename)

